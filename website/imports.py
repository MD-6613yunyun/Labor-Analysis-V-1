from flask import Blueprint,request,render_template,redirect,url_for,session, jsonify
from openpyxl import load_workbook
from website import db_connect, extract_shop_datas
from .views import get_data
from datetime import datetime,timedelta
from psycopg2.errors import IntegrityError
import re

imports = Blueprint('imports',__name__)

def get_partial_amount(percent,total):
    return round((float(percent)/100)*float(total),2)

def vehicle_plate_check(plate):
    conn = db_connect()
    cur = conn.cursor()
    if len(plate) not in (9,10):
        return f"Invalid Length of vehicle plate at Row "
    else:
        state_code = plate[:3]
        cur.execute("SELECT id,short_name FROM state;")
        state_codes = {data[1]:data[0] for data in cur.fetchall()}
        if state_code not in state_codes:
            return f"Invald State Code at Row "
        elif 'UN' in plate:
            return 1
        elif not plate[-4:].isdigit():
            return f"Invalid Last Digits Plate at Row "
        return 1


@imports.route("/excel",methods=['GET','POST'])
def excel_import():
    if 'pg_username' not in session or "pg_id" not in session:
        return redirect(url_for('views.home'))
    if request.method == 'POST':
        upload_file = request.files['upload_serivce_datas']
        excel_file_type = request.form.get('selectExcelFile')
        view_type = request.form.get('selectView')


        if upload_file.filename != '' and upload_file.filename.endswith(".xlsx"):
            workbook = load_workbook(filename=upload_file,data_only=True,read_only=True)
            try:
                worksheet = workbook[excel_file_type]
            except:
                return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f'The sheet name of excel import file must be <strong>{excel_file_type}..<&#47;strong>'))
            conn = db_connect()
            cur = conn.cursor()
            credentials , b_units = extract_shop_datas(cur)
            if excel_file_type == 'Jobs Data':
                eachJob_insert_query = """INSERT INTO eachJobline (form_id,job_name,job_type_id,total_amt,fst_pic_id,fst_pic_amt,sec_pic_id,sec_pic_amt,thrd_pic_id,thrd_pic_amt,frth_pic_id,
                frth_pic_amt,lst_pic_id,lst_pic_amt,line_concatenated,pic_rate_id) VALUES """
                cur.execute("SELECT id,name,shop_id FROM technicians;")
                technicians = cur.fetchall()
                technicians = {(data[1].lower(),data[2]):data[0] for data in technicians}
                cur.execute("SELECT id,name FROM jobType;")
                job_types = cur.fetchall()
                job_types = {data[1].lower():data[0] for data in job_types}
                cur.execute("SELECT fst_rate,sec_rate,thrd_rate,frth_rate,lst_rate,unique_rate,id FROM pic ORDER BY id;")
                all_rates = {data[5]:[data[:5],data[6]] for data in cur.fetchall()}
                cur.execute("SELECT shop.business_unit_id,shop.id as shop_id,LOWER(unit.name),LOWER(shop.name) FROM res_partner AS unit INNER JOIN shop ON shop.business_unit_id = unit.id;")
                unit_shop_datas = cur.fetchall()
                unit_shop_dct = {data[2:]:data[:2] for data in unit_shop_datas}
                conflict_unique_column = "line_concatenated"
                for row_counter,row in enumerate(worksheet.iter_rows(min_row=2),start=2):
                    # check invalid rate
                    if row[19].value is None or row[19].value.strip() == "" or row[19].value.strip() not in all_rates:
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid or Unregistered PIC Rate at row {row_counter}"))
                    # check all invalid fields
                    if None in (row[0].value,row[1].value,row[2].value,row[3].value,row[4].value,row[5].value,row[6].value,
                                row[8].value,row[9].value,row[10].value,row[11].value,row[17].value,row[18].value,row[19].value) or "" in (row[0].value,row[1].value,row[2].value,row[3].value,row[4].value,row[5].value,row[6].value,row[7].value,
                                row[8].value,row[9].value,row[10].value,row[11].value,row[17].value,row[18].value,row[19].value):
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid or Blank field at row {row_counter}"))
                    # get unit and shop id
                    unit_shop_ids = unit_shop_dct.get((row[17].value.strip().lower(),row[18].value.strip().lower()))
                    if not unit_shop_ids:
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid or Blank Business Unit at row {row_counter}"))
                    elif unit_shop_ids[1] not in [data[2] for data in credentials]:
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"You don't have enough access to import for {row[18].value.strip()}")) 
                    # get rate and assign tech name 
                    rate = all_rates[row[19].value.strip()]
                    cell_counter = 12
                    tech_names = []
                    for each_rate in rate[0]:
                        if each_rate:
                            if row[cell_counter].value is None or row[cell_counter].value.strip() == "":
                                return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Blank technician name at row {row_counter} and column {cell_counter+1}, must be matched with pic_rate..."))
                            if (row[cell_counter].value.strip().lower(),unit_shop_ids[1]) not in technicians:
                                return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid Technician Name at row - {row_counter} and column {cell_counter+1}.. \nPlease check or add names in Configurations -> Technicinans.."))
                            tech_names.append((row[cell_counter].value.strip().lower(),unit_shop_ids[1]))
                        else:
                            tech_names.append(('bot',None))
                        cell_counter += 1 
                    # get brand / model id
                    cur.execute("SELECT id,brand_id FROM vehicle_model WHERE LOWER(name) = %s;",(row[6].value.strip().lower(),))
                    idds = cur.fetchone()
                    if not idds:
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid or Unregistered Vehicle Model at row {row_counter}"))  
                    # check vehicle plate
                    returned_mgs = vehicle_plate_check(row[5].value.strip())
                    if returned_mgs != 1:
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"{returned_mgs} {row_counter}."))
                    # create vehicle
                    vehicle_year = row[7].value
                    # check vehicle
                    plate = row[5].value.strip() if 'UN' in row[5].value.strip() else row[5].value.strip().replace("-","")
                    cur.execute("SELECT id FROM vehicle WHERE REPLACE(LOWER(plate),'-','') = %s;",(plate,))
                    vehicle_datas = cur.fetchone()                    
                    phone = re.sub(r'\D', '', str(row[4].value).strip().split(",")[0])
                    if not phone.startswith("0"):
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid Phone {phone} at {row_counter}.\n Please Starts  customer phone with 09.."))   
                    cur.execute("SELECT id,name,phone,secondary_phone FROM customer WHERE phone = %s or secondary_phone = %s or name = %s;",(phone,phone,row[3].value.strip().upper()))
                    cus_datas = cur.fetchall()
                    if cus_datas:
                        for cus_data in cus_datas:
                            if cus_data[2] == phone or cus_data[3] == phone:
                                cus_id = cus_data[0]
                            if vehicle_datas:
                                cur.execute("SELECT customer_id FROM ownership WHERE vehicle_id = %s;",(vehicle_datas))
                                ownership_datas_for_vehicle_data = cur.fetchall()
                                for ownership_data in ownership_datas_for_vehicle_data:
                                    if ownership_data[0] == cus_data[0]:
                                        cur.execute("UPDATE customer SET secondary_phone = %s WHERE id = %s;",(phone,ownership_data[0]))
                                        cus_id = ownership_data[0]
                                        break
                            if cus_id:
                                break
                        if not cus_id:                       
                            cur.execute(""" INSERT INTO customer (name,phone)
                                    VALUES (%s,%s)
                                    ON CONFLICT (phone) DO NOTHING 
                                    RETURNING id """,(row[3].value.strip().upper(),phone))
                            cus_id = cur.fetchone()[0] 
                    else:
                        # create new customer
                        cur.execute(""" INSERT INTO customer (name,phone)
                                        VALUES (%s,%s)
                                        ON CONFLICT (phone) DO NOTHING 
                                        RETURNING id """,(row[3].value.strip().upper(),phone))
                        cus_id = cur.fetchone()[0]
                    if vehicle_datas:
                        vehicle_id = vehicle_datas[0]
                    else:
                        print(plate)
                        cur.execute(""" INSERT INTO vehicle (plate,model_id,brand_id,year)
                                        VALUES (%s,%s,%s,%s)
                                        ON CONFLICT (plate) DO UPDATE
                                        SET plate = EXCLUDED.plate
                                        RETURNING id;""",(plate.upper(),idds[0],idds[1],vehicle_year))
                        vehicle_id = cur.fetchone()[0]
                    if not cus_id:
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid Data for Customer ({row[3].value.strip()}-{phone}) and Vehicle ({row[5].value.strip()}) at -  {row_counter} "))            
                    # create ownership
                    cur.execute(""" WITH inserted AS (
                                        INSERT INTO ownership (vehicle_id,customer_id,unique_owner)
                                        VALUES (%s,%s,%s)
                                        ON CONFLICT (unique_owner) DO NOTHING 
                                        RETURNING id
                                    )
                                    SELECT id FROM inserted
                                    UNION ALL
                                    SELECT id FROM ownership WHERE unique_owner = %s 
                                    LIMIT 1;""",(vehicle_id,cus_id,f"{vehicle_id}-{cus_id}",f"{vehicle_id}-{cus_id}"))
                    ownership_id = cur.fetchall()[0][0]
                    # insert data into form
                    cur.execute(""" SELECT id FROM eachJObFORM WHERE job_date = %s AND shop_id = %s AND job_no = %s;""",(row[1].value.strftime("%Y-%m-%d"),str(unit_shop_ids[1]),row[2].value))
                    form_id = cur.fetchone()
                    if not form_id:
                        cur.execute(""" INSERT INTO eachJobForm (job_date,job_no,business_unit_id,shop_id,invoice_no,customer_id,vehicle_id,ownership_id) 
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s) RETURNING id;""",(row[1].value.strftime("%Y-%m-%d"),row[2].value,str(unit_shop_ids[0]),str(unit_shop_ids[1]),row[8].value,cus_id,vehicle_id,ownership_id))
                        form_id = cur.fetchone()[0]
                    else:
                        form_id = form_id[0]
                    # create psfu 
                    cur.execute("INSERT INTO psfu_call (form_id) VALUES (%s);",(form_id,))
                    # check job type
                    if row[10].value.strip().lower() not in job_types:
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid Job Type at row -  {row_counter}"))                    
                    eachJob_concatenated = row[1].value.strftime("%Y/%m/%d") + row[2].value + row[9].value + str(unit_shop_ids[1])
                    # extends query
                    try:
                        eachJob_insert_query += f"""('{form_id}','{row[9].value}','{job_types[row[10].value.strip().lower()]}','{int(row[11].value):.2f}','{technicians[tech_names[0]]}','{get_partial_amount(rate[0][0],row[11].value)}','{technicians[tech_names[1]]}','{get_partial_amount(rate[0][1],row[11].value)}','{technicians[tech_names[2]]}','{get_partial_amount(rate[0][2],row[11].value)}','{technicians[tech_names[3]]}','{get_partial_amount(rate[0][3],row[11].value)}','{technicians[tech_names[4]]}','{get_partial_amount(rate[0][4],row[11].value)}','{eachJob_concatenated}','{rate[1]}'),"""
                    except ValueError:
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid Job Data at row -  {row_counter}"))
                cur.execute(eachJob_insert_query[:-1] + f" ON CONFLICT ({conflict_unique_column}) DO NOTHING;")
                print(eachJob_insert_query[:-1])
                conn.commit()    
                cur.close()
                conn.close()
                return redirect(url_for('views.show_service_datas',typ='service-datas'))            
            elif excel_file_type == 'Types Data':
                eachJob_insert_query = "INSERT INTO jobType (name) VALUES "
                conflict_unique_column = "name"
                for row_counter,row in enumerate(worksheet.iter_rows(min_row=2),start=2):
                    eachJob_insert_query += f"('{row[0].value}'),"
            elif excel_file_type == 'Technicians': 
                eachJob_insert_query = "INSERT INTO technicians (name,business_unit_id,shop_id) VALUES "
                conflict_unique_column = "name"
                cur.execute("SELECT name,business_unit_id,id FROM shop;")
                shops_dct = {data[0]:data[1:] for data in cur.fetchall()}
                for row_counter,row in enumerate(worksheet.iter_rows(min_row=2),start=2):
                    shop_name = row[2].value.strip()
                    if None in (row[0].value,row[1].value,row[2].value) or row[0].value.strip() == "" or row[1].value.strip() == "" or shop_name == "" or shop_name not in shops_dct:
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid or Blank field at row {row_counter}"))                           
                    eachJob_insert_query += f"('{row[0].value.strip().upper()}','{shops_dct[shop_name][0]}','{shops_dct[shop_name][1]}'),"
            elif excel_file_type == 'brand-model':
                eachJob_insert_query = "INSERT INTO vehicle_model(brand_id,name) VALUES"
                for row_counter,row in enumerate(worksheet.iter_rows(min_row=2),start=2):
                    if None in (row[0].value,row[1].value) or row[0].value.strip() == "" or row[1].value.strip() == "":
                        return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f"Invalid or Blank field at row {row_counter}"))        
                    cur.execute(f"INSERT INTO vehicle_brand (name) VALUES('{row[0].value.strip()}') ON CONFLICT (name) DO UPDATE set name = '{row[0].value.strip()}' RETURNING id;")
                    eachJob_insert_query += f"({cur.fetchall()[0][0]},'{row[1].value}'),"
                conflict_unique_column = "name"
            cur.execute(eachJob_insert_query[:-1] + f" ON CONFLICT ({conflict_unique_column}) DO NOTHING;")
            conn.commit()
            cur.close()
            conn.close()
        else:
            return redirect(url_for('views.show_service_datas',typ=view_type,mgs=f'<strong> Invalid Excel File Type.. <&#47;strong>'))
    return redirect(url_for('views.show_service_datas',typ=view_type))

@imports.route("/create-form/<typ>")
def show_create_form(typ,mgs=None):
    if 'pg_username' not in session or "pg_id" not in session:
        return redirect(url_for('views.home'))
    mgs = request.args.get("mgs")
    conn = db_connect()
    cur = conn.cursor()
    result = ["","",""]
    credentials , b_units = extract_shop_datas(cur)
    if typ == 'service-datas':
        cur.execute(""" SELECT tech.id,tech.name FROM technicians tech
                            INNER JOIN (
                                SELECT technician_id, to_shop_id FROM technicians_transfer
                                WHERE 
                                    CURRENT_DATE BETWEEN from_date AND to_date
                                            OR
                                    CURRENT_DATE >= from_date AND to_date IS NULL 
                            ) AS transfer
                        ON transfer.technician_id = tech.id
                        WHERE tech.id != 0 AND transfer.to_shop_id = %s;""",(credentials[0][2],))
        result.append(cur.fetchall())
        cur.execute("SELECT id,name FROM jobType;")
        result.append(cur.fetchall())
        cur.execute("SELECT unique_rate FROM pic;")
        result.append(cur.fetchall())
        cur.execute("SELECT id,short_name,name FROM state;")
        result.append(cur.fetchall())
        cur.execute("SELECT id,name FROM vehicle_brand;")
        result.append(cur.fetchall())
        result.append(datetime.now().year)
        cur.execute("SELECT name FROM customer;")
        result.append(cur.fetchall())
    elif typ == 'check-in-out':
        result.append([])
        cur.execute("SELECT id,name FROM jobType;")
        result.append(cur.fetchall())
        cur.execute("SELECT id,short_name,name FROM state;")
        result.append(cur.fetchall())
        cur.execute("SELECT id,name FROM vehicle_brand;")
        result.append(cur.fetchall())
        result.append(datetime.now().year)
        cur.execute("SELECT id,name FROM descriptions;")
        result.append(cur.fetchall())
        cur.execute("SELECT 'JC/' || RIGHT(EXTRACT(YEAR FROM NOW())::TEXT,2) || '/' || TO_CHAR(EXTRACT(MONTH FROM NOW()),'FM00') || '/' || COALESCE((SELECT TO_CHAR(RIGHT(seq, 2)::INT + 1, 'FM0000') FROM checkinoutform ORDER BY seq DESC LIMIT 1),'0001');")
        result.append(cur.fetchone()[0])
    elif typ == 'customers-create':
        result = [datetime.now().strftime("%Y-%m-%d")]
        cur.execute("SELECT 'CUS'||LPAD((id+1)::text,7,'0') FROM customer ORDER BY id DESC LIMIT 1;")
        try:
            result.append(cur.fetchall()[0][0])
        except IndexError:
            result.append('CUS0000001')
        cur.execute("SELECT id,name FROM state;")
        result.append(cur.fetchall())
        return render_template('registration_form.html',result=result,typ=typ,credentials=credentials,b_units=b_units)
    elif typ == 'vehicles-create':
        result = [datetime.now().strftime("%Y-%m-%d")]
        cur.execute("SELECT name FROM vehicle_brand;")
        result.append(cur.fetchall())
        cur.execute("SELECT short_name FROM state;")
        result.append(cur.fetchall())
        result.append(datetime.now().year)
        return render_template('registration_form.html',result=result,typ=typ,credentials=credentials,b_units=b_units)  
    elif typ == 'leaves':
        cur.execute("SELECT id,name FROM leave_type;")
        result.append(cur.fetchall())
        print(result)

    return render_template('input_form.html',result=result,mgs=mgs,typ=typ,credentials=credentials,b_units=b_units)

@imports.route("/keep-in-import/<typ>",methods=['POST'])
def keep_in_import(typ): 
    if 'pg_username' not in session or "pg_id" not in session:
        return redirect(url_for('views.home'))       
    mgs = None
    if request.method == 'POST':
        conn = db_connect()
        cur = conn.cursor()
        credentials, b_units = extract_shop_datas(cur)
        if typ == 'service-datas':
            conflict_unique_column = 'line_concatenated'
            eachJob_insert_query = """INSERT INTO eachJobline (form_id,job_name,job_type_id,total_amt,fst_pic_id,fst_pic_amt,sec_pic_id,sec_pic_amt,thrd_pic_id,thrd_pic_amt,frth_pic_id,
                frth_pic_amt,lst_pic_id,lst_pic_amt,line_concatenated,pic_rate_id) VALUES """
            edit = request.form.get("newOrEdit")
            #
            job_no = request.form.get("jobNo")
            invoice_no = request.form.get("invoiceNo")
            shop_id = request.form.get("shop")
            job_date = request.form.get("jobDate")
            # getting
            cur.execute(""" SELECT id FROM eachJObFORM WHERE  (job_date = %s AND shop_id = %s AND job_no = %s) OR (job_date = %s AND shop_id = %s AND invoice_no = %s);""",(job_date,shop_id,job_no,job_date,shop_id,invoice_no))
            if cur.fetchone() != None and not edit:
                mgs = 'Job No. / Invoice No. is already existed in our system...'
            else:
                cur.execute("SELECT business_unit_id FROM shop WHERE id = %s;",(shop_id,))
                unit_id = cur.fetchall()[0][0]
                #
                descriptions = request.form.getlist("description")[1:]
                job_types = [data.strip() for data in request.form.getlist('jobType')[1:]]
                cur.execute("SELECT fst_rate,sec_rate,thrd_rate,frth_rate,lst_rate,unique_rate,id FROM pic;")
                all_rates = {data[5]:[data[:5],data[6]] for data in cur.fetchall()}
                cur.execute("SELECT id,name FROM technicians;")
                all_technicians = {data[1]:data[0] for data in cur.fetchall()}
                #
                pic_ones = [data for data in request.form.getlist("picOne")[1:] ]
                pic_twos = [data for data in request.form.getlist("picTwo")[1:] ]
                pic_threes = [data for data in request.form.getlist("picThree")[1:] ]
                pic_fours = [data for data in request.form.getlist("picFour")[1:] ]
                pic_fives = [data for data in request.form.getlist("picFive")[1:] ]
                pic_rates = [data.split(",") for data in request.form.getlist("pic-rate")]
                job_costs = request.form.getlist("jobCost")[1:]
                if edit:
                    old_job_no = request.form.get("oldJobNo")
                    print(get_data('eachJobDelForm',old_job_no))
                vehicle_id = request.form.get("vehicleInformation")
                new_onwership = False
                if request.form.get("vehicleInformation") == "None":
                    plate = request.form.get("regState") + request.form.get("regPrefix") + request.form.get("regDigits")
                    model = request.form.get("model_id")
                    brand = request.form.get("brand_id")
                    year = request.form.get("year")
                    cur.execute("INSERT INTO vehicle (plate,model_id,brand_id,year) VALUES (%s,%s,%s,%s) RETURNING id;",(plate,model,brand,year)) 
                    vehicle_id = cur.fetchall()[0][0]     
                    new_onwership = True    
                customer_id = request.form.get("customerInformation")
                if request.form.get("customerInformation") == "None":
                    cus_name = request.form.get("customerName")
                    address = request.form.get("fullAddress")
                    state_id = request.form.get("state")
                    phone = request.form.get("phone")
                    cur.execute("INSERT INTO customer (name,address,phone,state_id) VALUES (%s,%s,%s,%s) ON CONFLICT (phone) DO NOTHING RETURNING id;",(cus_name,address,phone,state_id))
                    customer_datas = cur.fetchall()
                    if customer_datas:
                        customer_id = customer_datas[0][0]
                        new_onwership = True
                    else:
                        return redirect(url_for('imports.show_create_form',typ='service-datas',mgs='Customer with same phone number is existed..'))
                if new_onwership:
                    cur.execute("INSERT INTO ownership (customer_id,vehicle_id,start_date,unique_owner) VALUES (%s,%s,%s,%s) RETURNING id;",(customer_id,vehicle_id,datetime.now().strftime("%Y-%m-%d"),str(vehicle_id)+'-'+str(customer_id)))
                else:
                    cur.execute("SELECT id FROM ownership WHERE customer_id = %s and vehicle_id = %s;",(customer_id,vehicle_id))                
                ownership_id = cur.fetchall()[0][0]
                # inserting form
                cur.execute("""INSERT INTO eachJobForm (job_date,job_no,business_unit_id,shop_id,invoice_no,
                    customer_id,vehicle_id,ownership_id) VALUES (%s, %s, %s, %s, %s, %s, %s, %s) RETURNING id;""",(job_date,job_no,unit_id,shop_id,invoice_no,customer_id,vehicle_id,ownership_id))
                form_id = cur.fetchone()[0]
                # create psfu 
                cur.execute("INSERT INTO psfu_call (form_id) VALUES (%s);",(form_id,))
                for data in zip(descriptions,job_types,pic_ones,pic_twos,pic_threes,pic_fours,pic_fives,job_costs,pic_rates):
                    eachJob_concatenated = job_date.replace("-","/") + job_no + data[0] + shop_id
                    eachJob_insert_query += f"""('{form_id}','{data[0]}','{data[1]}','{data[7]}','{all_technicians.get(data[2],0)}',
                    '{get_partial_amount(data[8][0],data[7])}','{all_technicians.get(data[3],0)}','{get_partial_amount(data[8][1],data[7])}',
                    '{all_technicians.get(data[4],0)}','{get_partial_amount(data[8][2],data[7])}','{all_technicians.get(data[5],0)}','{get_partial_amount(data[8][3],data[7])}',
                    '{all_technicians.get(data[6],0)}','{get_partial_amount(data[8][4],data[7])}','{eachJob_concatenated}','{all_rates[",".join(data[8])][1]}'),"""
                cur.execute(eachJob_insert_query[:-1] + f" ON CONFLICT ({conflict_unique_column}) DO NOTHING;")
                conn.commit()
            return redirect(url_for('imports.show_create_form',typ='service-datas',mgs=mgs))
        elif typ == 'check-in-out-techs':
            client_data = request.json
            date_value = client_data['job_date']
            tech_id = client_data['tech_id']
            start_time = client_data['start_time']
            end_time = client_data['end_time']
            cur.execute(""" SELECT	line.technician_id
                                FROM checkinoutform AS form
                            INNER JOIN checkinoutline AS line
                                ON form.id = line.form_id
                            WHERE form.job_date = %s AND technician_id = %s AND 
                                (
                                    (%s BETWEEN line.start_time AND line.end_time)
                                            OR
                                    (%s BETWEEN line.start_time AND line.end_time)
                                            OR
                                    (%s <= line.start_time AND %s >= line.end_time)
                                );""",(date_value,tech_id,start_time+':00',end_time+':00',start_time+':00',end_time+':00'))
            if not cur.fetchone():
                res = [1]
            else:
                res = [0]
            return jsonify(res)
        elif typ == 'check-in-out':
            vehicle_id = request.form.get("vehicleInformation")
            job_no = request.form.get("jobNo")
            job_date = request.form.get("jobDate")
            shop_id = request.form.get("shop")
            approve = request.form.get("approve")
            edit = request.form.get("newOrEdit")
            status = request.form.get("status")
            form_seq = request.form.get("form_seq")
            no_error = False

            print(approve)
            print(edit,"hel")
            if approve:
                cur.execute("SELECT id FROM checkInOutForm WHERE (job_no = %s and job_date = %s and shop_id = %s);",(job_no,job_date,shop_id))
                time_form_job_no = cur.fetchall()
                if job_no == "":
                    mgs = "Invalid Empty Job Number...."
                elif time_form_job_no != None and len(time_form_job_no) > 1:
                    mgs = f'Job No. is already existed with date ( {job_date} ) in our system...'
                else:
                    today_date = datetime.strptime(job_date, '%Y-%m-%d').date()
                    cur.execute("SELECT id FROM eachJobForm WHERE shop_id = %s AND job_no = %s AND job_date BETWEEN %s  AND %s;",(shop_id,job_no,today_date - timedelta(days=30), today_date + timedelta(days=30)))
                    job_form_id = cur.fetchone()
                    if not job_form_id:
                        mgs = "There is no matching job no in Servcie Datas.."
                    else:
                        cur.execute("UPDATE checkinoutform SET form_id = %s WHERE id = %s;",(job_form_id,approve))
                    conn.commit()
                no_error = True
                form_id = approve
            else:
                if vehicle_id == "None":
                    plate = request.form.get("regState") + request.form.get("regPrefix") + request.form.get("regDigits")
                    model = request.form.get("model_id")
                    brand = request.form.get("brand_id")
                    year = request.form.get("year")
                    cur.execute("INSERT INTO vehicle (plate,model_id,brand_id,year) VALUES (%s,%s,%s,%s) RETURNING id;",(plate,model,brand,year)) 
                    vehicle_id = cur.fetchone()[0] 
                check_in_out_query = "INSERT INTO checkInOutLine (form_id,description_id,technician_id,job_type_id,start_time,end_time,line_concatenated) VALUES "
                cur.execute("SELECT business_unit_id FROM shop WHERE id = %s;",(shop_id,))
                unit_id = cur.fetchall()[0][0]
                try:
                    if status == "None":
                        cur.execute("INSERT INTO checkInOutForm (seq,vehicle_id,job_date,job_no,business_unit_id,shop_id) VALUES (%s,%s,%s,%s,%s,%s) RETURNING id;",(form_seq,vehicle_id,job_date,job_no,unit_id,shop_id))
                    else:
                        cur.execute("INSERT INTO checkInOutForm (seq,vehicle_id,job_date,job_no,business_unit_id,shop_id,form_id) VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id;",(form_seq,vehicle_id,job_date,job_no,unit_id,shop_id,status))
                    form_id = cur.fetchone()[0]
                except IntegrityError as err:
                    print(err)
                    mgs = "Error while inserting form data into database..."
                if not mgs:
                    #
                    descriptions = request.form.getlist("description_id")[1:]
                    job_types = [data.strip() for data in request.form.getlist('jobType')[1:]]
                    technicians = request.form.getlist("pic_id")[1:]
                    start_times = request.form.getlist("startTimeVal")[1:]
                    end_times = request.form.getlist("endTimeVal")[1:]

                    for data in zip(descriptions, technicians, job_types, start_times, end_times):
                        line_concat = f"{data[0]},{data[1].strip()},{data[2]},{job_date},{data[3]},{data[4]},{form_id}"
                        check_in_out_query += f"({form_id},'{data[0]}',{data[1].strip()},{data[2]},'{data[3]}','{data[4]}','{line_concat}'),"
                    try:
                        print(check_in_out_query[:-1] + " ON CONFLICT (line_concatenated) DO NOTHING;")
                        cur.execute(check_in_out_query[:-1] + " ON CONFLICT (line_concatenated) DO NOTHING;")
                        if edit:
                            print(get_data('checkinoutDelForm',edit))
                        conn.commit()
                    except IntegrityError as err:
                        print(err)
                        conn.rollback()
                        mgs = "ERROR: Unknown error occurred while inserting data into database.."
                        return redirect(url_for('imports.show_create_form',typ='check-in-out',mgs=mgs))
                    else:
                        no_error = True
            if no_error:
                with_id_query =""" SELECT car.plate , car.id ,  form.job_date , form.job_no , brand.name , model.name , brand.id , model.id , car.year, unit.name , shop.name , form.id , 
                                        line.id , descriptions.name , tech.name , type.name , line.start_time , line.end_time , line.end_time - line.start_time , shop.id , form.form_id , line.description_id , tech.id , form.seq
                                        FROM checkinoutform AS form
                                        INNER JOIN checkinoutline AS line
                                        ON form.id = line.form_id
                                        LEFT JOIN descriptions
                                        ON descriptions.id = line.description_id
                                        LEFT JOIN vehicle AS car
                                        ON car.id = form.vehicle_id
                                        LEFT JOIN vehicle_brand AS brand
                                        ON brand.id = car.brand_id
                                        LEFT JOIN vehicle_model AS model
                                        ON model.id = car.model_id
                                        LEFT JOIN res_partner AS unit
                                        ON unit.id = form.business_unit_id
                                        LEFT JOIN shop AS shop
                                        ON shop.id = form.shop_id
                                        LEFT JOIN technicians AS tech
                                        ON tech.id = line.technician_id
                                        LEFT JOIN jobType AS type
                                        ON type.id = line.job_type_id
                                        WHERE form.id = %s
                                        ORDER BY line.start_time;""" 
                print(form_id)
                result = []
                cur.execute("SELECT id,name FROM technicians WHERE shop_id = %s;",(credentials[0][2],))
                result.append(cur.fetchall())
                cur.execute("SELECT id,name FROM jobType;")
                result.append(cur.fetchall())
                cur.execute("SELECT id,short_name,name FROM state;")
                result.append(cur.fetchall())
                cur.execute("SELECT id,name FROM vehicle_brand;")
                result.append(cur.fetchall())
                result.append(datetime.now().year)
                cur.execute(with_id_query,(form_id,))
                result.append(cur.fetchall())
                result.append(sum([dt[18] for dt in result[5]],timedelta()))
                cur.execute("SELECT id,name FROM descriptions;")
                result.append(cur.fetchall())
                return render_template('edit_form.html',result = result,typ=typ,mgs=mgs,credentials=credentials,b_units=b_units)
            return redirect(url_for('imports.show_create_form',typ='check-in-out',mgs=mgs))
        elif typ == 'leaves':
            leave_type_id = request.form.get("leave-type-id")
            shop_id = request.form.get("shop")
            techs = request.form.getlist("tech")
            start_dt = request.form.get("startDate")
            duration = request.form.get("duration")
            end_dt = request.form.get("endDate")
            remark = request.form.get("remark")

            if datetime.strptime(end_dt, '%Y-%m-%d') < datetime.strptime(start_dt, '%Y-%m-%d'):
                return redirect(url_for('imports.show_create_form',typ=typ,mgs="End Date must be greater than  start_date.."))
            elif start_dt == end_dt:
                if duration == '1':
                    start_dt += ' 00:00:00'
                    end_dt += ' 24:00:00'
                elif duration == '2':
                    start_dt += ' 00:00:00'
                    end_dt += ' 12:00:00'
                elif duration == '3':
                    start_dt += ' 12:00:00'
                    end_dt += ' 24:00:00'


            if leave_type_id:
                # for people
                for tech_id in techs[1:]:
                    cur.execute("""INSERT INTO leaves(leave_type_id,shop_id,technician_id,start_date,end_date,remark) 
                                SELECT %s,%s,%s,%s,%s,%s
                                WHERE NOT EXISTS (SELECT 1 FROM leaves WHERE leave_type_id = %s AND shop_id = %s  AND technician_id = %s AND
                                start_date = %s AND end_date = %s AND remark = %s);""",(leave_type_id,shop_id,tech_id,start_dt,end_dt,remark,leave_type_id,shop_id,tech_id,start_dt,end_dt,remark))
                
            else:
                # for shops
                for shop_id in techs[1:]:
                    cur.execute(""" INSERT INTO shop_leaves(shop_id,start_date,end_date,remark) 
                                SELECT %s,%s,%s,%s
                                WHERE NOT EXISTS (SELECT 1 FROM shop_leaves WHERE shop_id = %s AND start_date = %s AND end_date = %s and remark = %s);"""
                                ,(shop_id,start_dt,end_dt,remark,shop_id,start_dt,end_dt,remark))
                    
                    cur.execute(""" INSERT INTO leaves(leave_type_id,shop_id,technician_id,start_date,end_date,remark)
                                        SELECT (SELECT id FROM leave_type WHERE name = %s),%s,tech.id,%s,%s,%s FROM technicians tech
                                        INNER JOIN (
                                            SELECT technician_id, to_shop_id FROM technicians_transfer
                                            WHERE 
                                                %s BETWEEN from_date AND to_date
                                                        OR
                                                %s >= from_date AND to_date IS NULL 
                                        ) AS transfer
                                    ON transfer.technician_id = tech.id
                                        INNER JOIN shop 
                                    ON shop.id = transfer.to_shop_id
                                    WHERE tech.id != 0 AND shop.id = %s ORDER BY tech.id;""",
                                    ('WHOLE SHOP LEAVE',shop_id,start_dt,end_dt,remark,start_dt,start_dt,shop_id))
            print(leave_type_id)
            print(shop_id)
            print(duration)
            print(techs)
            print(start_dt)
            print(end_dt)
            print(remark)
            conn.commit()
            return redirect(url_for('imports.show_create_form',typ=typ))
        elif typ == 'pic-rate':
            idd = request.form.get("idd")
            rates = request.form.getlist('rate')
            if idd:
                rates = rates[5:]
            rates = [rate.strip() for rate in rates]
            rates.append(','.join(rates))
            if not idd:
                cur.execute(f"INSERT INTO pic (fst_rate,sec_rate,thrd_rate,frth_rate,lst_rate,unique_rate) VALUES {tuple(rates)} ON CONFLICT (unique_rate) DO NOTHING;")
            else:
                cur.execute(f""" UPDATE pic SET fst_rate = {rates[0]},sec_rate = {rates[1]},thrd_rate = {rates[2]},frth_rate = {rates[3]},lst_rate = {rates[4]},unique_rate = '{rates[5]}' WHERE id = {idd} AND NOT EXISTS (SELECT 1 FROM pic WHERE unique_rate = '{rates[5]}'); """)
        elif typ == 'psfu-calls':
            psfu_id = request.form.get('psfu_id')
            call_status_id = request.form.get('call_status_id')
            remark = request.form.get('remark')
            cur.execute("UPDATE psfu_call SET call_status_id = %s , remark = %s WHERE id = %s;",(call_status_id,remark,psfu_id))
            conn.commit()
        elif typ == 'technician':
            tech_name  = request.form.get("tech").strip()
            tech_id = request.form.get("tech_id")
            edit_form = False
            if tech_id:
                if tech_name:
                    cur.execute("UPDATE technicians SET name = %s WHERE id = %s AND NOT EXISTS (SELECT 1 FROM technicians WHERE name = %s) RETURNING id;",(tech_name,tech_id,tech_name))
                    if not cur.fetchone():
                        mgs = "Name is already Existed..."
                else:
                    from_shop = request.form.get("from-shop")
                    to_shop = request.form.get("to-shop")
                    start_dt = request.form.get("start-dt")
                    remark = request.form.get("remark")
                    cur.execute("""UPDATE technicians_transfer SET 
                                to_date = to_date(%s, 'YYYY-MM-DD') - INTERVAL '1 days'
                                WHERE technician_id = %s AND to_date IS NULL;""",(start_dt,tech_id))
                    cur.execute("""INSERT INTO technicians_transfer (technician_id, from_shop_id, to_shop_id, from_date, remark, created_by)
                                    VALUES (%s, %s, %s, %s, %s, %s);""",(tech_id, from_shop, to_shop, start_dt, remark, credentials[0][0]))
                conn.commit()
                edit_form = True

            if edit_form:
                cur.execute(""" SELECT his.technician_id, tech.name, from_shop.name , to_shop.id ,to_shop.name, his.from_date, 
                                his.to_date,CASE
		                                        WHEN his.to_date IS NULL THEN 1
							                    WHEN CURRENT_DATE BETWEEN his.from_date AND his.to_date THEN 0
		                                        ELSE 2
                                            END AS flag_shop
                                    FROM technicians_transfer AS his
                                LEFT JOIN technicians AS tech
                                    ON tech.id = his.technician_id
                                LEFT JOIN shop AS from_shop
                                    ON from_shop.id = his.from_shop_id
                                LEFT JOIN shop AS to_shop
                                    ON to_shop.id = his.to_shop_id
                                WHERE tech.id = %s
				                    ORDER BY flag_shop;""",(tech_id,))
                result = []
                result.append(cur.fetchall())
                cur.execute("SELECT id,name FROM shop;")
                result.append(cur.fetchall())
                return render_template('edit_form.html',result = result,typ=typ, credentials = credentials ,b_units = b_units, mgs=mgs)
            
            shop_name = request.form.get("shop")
            joined_date = request.form.get("joined-date")
            cur.execute("SELECT business_unit_id,id FROM shop WHERE name = %s;",(shop_name,))
            data = cur.fetchone()
            if not data:
                mgs = 'Invalid Shop / Business Unit'
            else:
                cur.execute(""" INSERT INTO technicians (name,business_unit_id,shop_id) VALUES (%s,%s,%s) ON CONFLICT (name) DO NOTHING RETURNING id;""",(tech_name,data[0],data[1]))
                tech_id = cur.fetchone()
                if not tech_id:
                    mgs = 'Technician Name is already existed....'
                else:
                    cur.execute(""" INSERT INTO  technicians_transfer (technician_id,from_shop_id,
                                to_shop_id,from_date,created_by) VALUES (%s,%s,%s,%s,%s);""",(tech_id,data[1],data[1],joined_date,credentials[0][0]))
        elif typ == 'jobType' or typ == 'descriptions' or typ == 'leave-type' or typ == 'call-status':
            print(typ)
            idd = request.form.get("idd")
            job_type = request.form.getlist("jobType")
            naming_job_type = job_type[1] if idd else job_type[0]
            typ = typ.replace("-","_")
            print(idd)
            if naming_job_type.strip() == '':
                mgs = "Invalid Name..."
            else:
                if idd:
                    cur.execute(f""" UPDATE {typ} SET name = '{naming_job_type}' WHERE id = '{idd}' and NOT EXISTS ( SELECT 1 FROM {typ} WHERE name = '{naming_job_type}');""")
                else:
                    cur.execute(f""" INSERT INTO {typ} (name) VALUES ('{naming_job_type}') ON CONFLICT (name) DO NOTHING;""")
            typ = typ.replace("_","-")
        elif typ == 'customers':
            reg_date = request.form.get("register-date")
            name = request.form.get("name")
            state = request.form.get("state")
            address = request.form.get("address")
            phone = request.form.get("phone")
            vehicle_ids = request.form.getlist("vehicleIds")
            customer_id = request.form.get("customerId")
            unique_phone_query = f"AND id <> '{customer_id}'" if customer_id else ""
            cur.execute(f"SELECT id FROM customer WHERE phone = '{phone}' {unique_phone_query};")     
            if customer_id:
                if len(cur.fetchall()) == 0:
                    cur.execute("UPDATE customer SET name = %s,state_id = %s,address = %s,phone = %s,registered_date = %s WHERE id = %s;",(name,state,address,phone,reg_date,customer_id))
                else:
                    mgs = f"Phone - {phone} is already existed.."                    
                for vehicle_id in vehicle_ids:
                    cur.execute("INSERT INTO customer (vehicle_id,customer_id,unique_owner) VALUES (%s,%s,%s) ON CONFLICT (unique_owner) DO NOTHING;",(vehicle_id,customer_id,vehicle_id + '-' +customer_id))     
            else:
                if len(cur.fetchall()) == 0:
                    cur.execute("INSERT INTO customer (name,address,phone,state_id,registered_date) VALUES (%s,%s,%s,%s,%s);",(name,address,phone,state,reg_date))
                else:
                    mgs = f"Phone - {phone} is already existed.."
        elif typ == 'vehicles':
            vehicle_id = request.form.get("vehicle-id")
            reg_date = request.form.get("register-date")
            plate = request.form.get("fst-part") + request.form.get("sec-part") + request.form.get("thrd-part")
            brand_id = request.form.get("brand_id")
            model_id = request.form.get("model_id")
            year = request.form.get("year")
            unique_plate_query = f"AND id <> '{vehicle_id}'" if vehicle_id else ""
            cur.execute(f"SELECT id FROM vehicle WHERE plate = '{plate}' {unique_plate_query};")  
            if vehicle_id:
                if len(cur.fetchall()) == 0:
                    cur.execute("UPDATE vehicle SET register_date = %s,plate = %s,brand_id = %s,model_id = %s,year = %s WHERE id = %s;",(reg_date,plate,brand_id,model_id,year,vehicle_id))
                else:
                    mgs = f"Plate - {plate} is already registered in our system.."
            else:
                if len(cur.fetchall()) == 0:
                    cur.execute("INSERT INTO vehicle (plate,brand_id,model_id,year,register_date) VALUES (%s,%s,%s,%s,%s);",(plate,brand_id,model_id,year,reg_date))
                else:
                    mgs = f"Plate - {plate} is already registered in our system.."
        elif typ == 'brand':
            brand_name = request.form.get("brand").strip()
            model_name = request.form.get("model").strip()
            cur.execute("INSERT INTO vehicle_brand(name) VALUES(%s) ON CONFLICT (name) DO UPDATE SET name = %s RETURNING id;",(brand_name,brand_name))
            brand_id = cur.fetchall()[0][0]
            cur.execute("SELECT id FROM vehicle_model WHERE brand_id = %s AND name = %s;",(brand_id,model_name))
            if len(cur.fetchall()) == 0:
                cur.execute("INSERT INTO vehicle_model(brand_id,name) VALUES (%s,%s);",(brand_id,model_name))
        conn.commit()
    return redirect(url_for('views.show_service_datas',typ=typ,mgs=mgs))
